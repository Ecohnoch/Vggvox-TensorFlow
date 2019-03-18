import copy
import math
import random
import numpy as np 
import pandas as pd
from openpyxl import load_workbook


# TODO List
# http://scikit-learn.org/stable/modules/cross_validation.html#cross-validation-iterators-for-i-i-d-data
def cross_val_score():
	pass

def cross_val_predict():
	pass

def kfold():
	pass

def leaveOneOut():
	pass

def shuffleSplit():
	pass


# Read Txt and Xlsx files
class reader:
	def __init__(self, filePath, fileType='txt', headerGap='\t', dataGap='\t', headerEndWith=None, dataEndWith=None, dataType='number'):
		# Example: >>>obj = reader('data.txt', fileType='txt')
		self.filePath  = filePath
		self.fileType  = fileType
		self.headerGap = headerGap
		self.dataGap   = dataGap
		self.headerEndWith = headerEndWith
		self.dataEndWith   = dataEndWith
		self.dataType = dataType

		# Basic info
		self.data   = []
		self.header = []
		self.shape  = ()

	def read_file(self):
		if self.fileType == 'txt':
			with open(self.filePath, 'r+') as f:
				header_text = f.readline()
				if self.headerEndWith == None:
					self.header = header_text.rstrip('\n').split(self.headerGap)
				else:
					self.header = header_text.rstrip('\n').rstrip(self.headerEndWith).split(self.headerGap)
				for i in range(len(self.header)):
					self.header[i] = self.header[i].strip()
				for line in f:
					if self.dataEndWith == None:
						if self.dataType == 'number':
							self.data.append([float(x) for x in line.rstrip('\n').split(self.dataGap)])
						elif self.dataType == 'str':
							self.data.append([x for x in line.rstrip('\n').split(self.dataGap)])
					else:
						if self.dataType == 'number':
							self.data.append([float(x) for x in line.rstrip('\n').rstrip(self.dataEndWith.split(self.dataGap))])
						elif self.dataType == 'str':
							self.data.append([x for x in line.rstrip('\n').rstrip(self.dataEndWith.split(self.dataGap))])

		elif self.fileType == 'xlsx':
			wb = load_workbook(self.filePath)
			sheet = wb.active
			for item in sheet[1]:
				self.header.append(item.value)
			for i in range(sheet.max_row - 1):
				# i + 2
				line = []
				for j in sheet[i+2]:
					line.append(float(j.value))
				self.data.append(line)
		elif self.fileType == 'csv':
			all_data = pd.read_csv(self.filePath)
			keys = all_data.keys()
			for item in keys:
				self.header.append(item)
			for i in all_data.values:
				line = []
				for j in i:
					line.append(j)
				self.data.append(line)

		# Check data
		self.shape = (len(self.data), len(self.header))
		for i in self.data:
			if len(i) != self.shape[1]:
				print("*** Error with data missing, when shape is %d rows and %d cols but the data exists %d cols. " % (self.shape[0], self.shape[1], len(i)), i)
				return
		# Data read suc
		print('SUC: Read source data suc!')



	def print_header(self):
		counter = 1
		print('Header: ')
		for i in self.header:
			print('\tKey %d: ' % (counter), i)
			counter += 1

	def print_data(self):
		counter = 1
		print('Data: ')
		for i in self.data:
			print('\tLine %d: ' % (counter), i)
			counter += 1

	def choose_target(self, target):
		self.target = []
		tar_pos = -1
		for i in range(len(self.header)):
			if self.header[i] == target:
				tar_pos = i
				break
		if tar_pos == -1:
			print("*** Error with invalid target name, please retype the target.", tar_pos)
			return 
		# Separate the origin header and data
		# Now header is : header - target
		# Now data   is : data   - target_col
		self.header.pop(tar_pos)
		for i in range(len(self.data)):
			self.target.append(self.data[i][tar_pos])
			self.data[i].pop(tar_pos)
		self.shape = (len(self.data), len(self.header))

	def merge_data_target(self):
		ans = []
		for i in range(len(self.data)):
			ans_line = [self.target[i]]
			for j in range(len(self.data[i])):
				ans_line.append(self.data[i][j])
			ans.append(ans_line)
		return ans

	def split_data_target(self, all_data):
		data = []
		target = []
		for i in all_data:
			data_line = []
			for j in range(len(i)):
				if j == 0:
					target.append(i[j])
				else:
					data_line.append(i[j])
			data.append(data_line)
		return data, target



	def train_test_split(self, test_size=0.2, random_state=42, shuffle=True, y_binary=False):
		random.seed(random_state)
		self.X_train = []
		self.y_train = []
		self.X_test  = []
		self.y_test  = []
		self.y_binary = {}
		
		test_data_num = math.floor(self.shape[0] * test_size)

		if shuffle:
			all_data = self.merge_data_target()
			random.shuffle(all_data)
			self.X_train, self.y_train = self.split_data_target(all_data)

			self.X_test  = self.X_train[0:test_data_num]
			self.X_train = self.X_train[test_data_num:]
			self.y_test  = self.y_train[0:test_data_num]
			self.y_train  = self.y_train[test_data_num:]
		else:
			self.X_test  = self.X_train[0:test_data_num]
			self.X_train = self.X_train[test_data_num:]
			self.y_test  = self.y_train[0:test_data_num]
			self.y_train  = self.y_train[test_data_num:]
		# Finished
		print('SUC: Train(%d) & Test(%d) has been split suc. Call obj.print_X_train etc and try.'%(len(self.X_train), len(self.X_test)))

	def print_X_train(self):
		print('X_train (row %d, col %d)' % (len(self.X_train), self.shape[1]))
		counter = 1
		for i in self.X_train:
			print("\tRow %d: "%(counter), i)
			counter += 1
	def print_X_test(self):
		print('X_test (row %d, col %d)' % (len(self.X_test), self.shape[1]))
		counter = 1
		for i in self.X_test:
			print("\tRow %d: "%(counter), i)
			counter += 1
	def print_y_train(self):
		print('y_train (row %d)' % (len(self.y_train)))
		counter = 1
		for i in self.y_train:
			print("\tRow %d: "%(counter), i)
			counter += 1
	def print_y_test(self):
		print('y_test (row %d)' % (len(self.y_test)))
		counter = 1
		for i in self.y_test:
			print("\tRow %d: "%(counter), i)
			counter += 1


if __name__ == '__main__':
	# txt
	# rd = reader('../Test_DataSet/COfreewy/COfreewy.txt', headerGap='\t', dataGap='\t')
	# rd.read_file()
	# rd.choose_target('Wind')
	# rd.print_header()
	# rd.print_data()
	# print(rd.shape)
	# rd.train_test_split(shuffle=False)
	# rd.print_X_train()
	# rd.print_X_test()
	# rd.print_y_train()
	# rd.print_y_test()

	# xlsx
	rd = reader('./feature.csv', 'csv')
	rd.read_file()
	rd.choose_target('sentiment')
	rd.print_header()
	# print(rd.target)
	rd.train_test_split(shuffle=True)
	rd.print_y_train()